import os
import re
import glob
import datetime
import csv
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication

# Attempt to import optional analytics packages
try:
    import pandas as pd
except ImportError:
    pd = None

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import docx
except ImportError:
    docx = None


def get_current_date_str():
    return datetime.datetime.now().strftime("%Y-%m-%d")


def clean_text_whitespace(text):
    if not isinstance(text, str):
        return ""
    return " ".join(text.strip().split())


def read_docx_fallback(file_path):
    import zipfile
    import xml.etree.ElementTree as ET
    try:
        texts = []
        with zipfile.ZipFile(file_path) as z:
            xml_content = z.read('word/document.xml')
            root = ET.fromstring(xml_content)
            namespace = {'w': 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'}
            for para in root.findall('.//w:p', namespace):
                p_text = []
                for run in para.findall('.//w:t', namespace):
                    if run.text:
                        p_text.append(run.text)
                if p_text:
                    texts.append("".join(p_text))
        return "\n".join(texts)
    except Exception as e:
        return f"[Fallback Reader Error: {e}]"


def read_text_document(file_path):
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")
    ext = os.path.splitext(file_path)[1].lower()
    if ext == ".txt":
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()
    elif ext == ".docx":
        if docx:
            try:
                doc = docx.Document(file_path)
                return "\n".join([p.text for p in doc.paragraphs])
            except Exception:
                return read_docx_fallback(file_path)
        else:
            return read_docx_fallback(file_path)
    elif ext == ".doc":
        with open(file_path, "rb") as f:
            content = f.read()
        words = re.findall(rb'[\x20-\x7E\x80-\xFF]+', content)
        try:
            return b" ".join(words).decode("cp1251", errors="ignore")
        except Exception:
            return b" ".join(words).decode("utf-8", errors="ignore")
    else:
        raise ValueError(f"Unsupported document extension: {ext}")


# =====================================================================
# AGENT 1: DUPLICATES ANALYZER
# =====================================================================

class DuplicatesAnalyzerAgent:
    def __init__(self, workspace_dir):
        self.workspace_dir = workspace_dir
        self.title_parasites = ["курс", "тренинг", "вебинар", "онлайн", "программа"]
        self.desc_stop_words = ["курс", "обучение", "тренинг", "вы узнаете", "вы научитесь", "мы рассмотрим", "изучим", "разберемся"]
        self.audience_markers = ["детей", "подростков", "премиальных", "мсб", "корпоративных", "сотрудников", "для бизнеса", "для жизни"]
        self.level_basic = ["базовый", "введение", "основы", "для начинающих"]
        self.level_advanced = ["продвинутый", "углубленный", "для экспертов", "middle", "senior"]

    def clean_title(self, name):
        if not name or not isinstance(name, str):
            return ""
        name = name.lower()
        for word in self.title_parasites:
            name = re.sub(r'\b' + word + r'\b', '', name)
        name = name.replace(".", " ")
        return " ".join(name.split())

    def clean_description(self, desc):
        if not desc or not isinstance(desc, str):
            return set()
        desc = desc.lower()
        for sw in self.desc_stop_words:
            desc = desc.replace(sw, " ")
        words = re.findall(r'\b\w{4,}\b', desc)
        return set(words)

    def extract_quotes_content(self, text):
        if not text or not isinstance(text, str):
            return []
        quotes_regex = r'[«"\'](.*?)[»"\'"]'
        return [q.strip().lower() for q in re.findall(quotes_regex, text) if q.strip()]

    def get_audience_and_levels(self, name):
        name_lower = name.lower() if isinstance(name, str) else ""
        auds = [m for m in self.audience_markers if m in name_lower]
        lbl = None
        if any(b in name_lower for b in self.level_basic):
            lbl = "basic"
        elif any(a in name_lower for a in self.level_advanced):
            lbl = "advanced"
        return set(auds), lbl

    def analyze(self, file_name="Свод электронного контента.xlsx"):
        file_path = os.path.join(self.workspace_dir, file_name)
        if not os.path.exists(file_path):
            alternatives = glob.glob(os.path.join(self.workspace_dir, "**", file_name), recursive=True)
            if alternatives:
                file_path = alternatives[0]
            else:
                raise FileNotFoundError(f"Input file '{file_name}' not found under '{self.workspace_dir}'.")

        print(f"[DuplicatesAnalyzerAgent] Reading document: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()
        if pd:
            if ext == '.csv':
                df = pd.read_csv(file_path)
            elif ext == '.xlsx':
                df = pd.read_excel(file_path)
            else:
                raise ValueError(f"Unsupported file format: {ext}")
        else:
            if ext == '.xlsx' and openpyxl:
                wb = openpyxl.load_workbook(file_path, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                hdr = rows[0]
                df_dict = {col: [] for col in hdr if col}
                for row in rows[1:]:
                    for idx, val in enumerate(row):
                        if idx < len(hdr) and hdr[idx]:
                            df_dict[hdr[idx]].append(val)
                df = df_dict
            else:
                raise ValueError("Cannot read file without pandas/openpyxl.")

        keys = list(df.keys()) if hasattr(df, 'keys') else list(df.keys())
        id_col = keys[1]
        title_col = keys[2] if len(keys) > 2 else next((c for c in keys if "назва" in c.lower()), keys[0])
        desc_col = next((c for c in keys if "опис" in c.lower()), keys[-1] if len(keys) > 3 else keys[0])

        courses = []
        num_rows = len(df[id_col])
        for i in range(num_rows):
            course_id = df[id_col][i]
            title = df[title_col][i]
            desc = df[desc_col][i]
            if not course_id or str(course_id).strip() == "nan" or not title:
                continue
            courses.append({
                "raw_id": str(course_id).strip(),
                "raw_title": str(title).strip(),
                "raw_desc": str(desc).strip() if desc and str(desc).strip() != "nan" else "",
                "clean_title": self.clean_title(str(title)),
                "clean_desc_words": self.clean_description(str(desc)) if desc else set()
            })

        print(f"[DuplicatesAnalyzerAgent] Loaded {len(courses)} courses.")
        matches = []

        for i, c1 in enumerate(courses):
            matched_pair = None
            match_type = None
            comment = "—"
            status = "Требует проверки"
            matched_ids = []

            for j, c2 in enumerate(courses):
                if i == j:
                    continue
                if c1["clean_title"] == c2["clean_title"] and c1["clean_title"] != "":
                    matched_ids.append(c2["raw_id"])
                    if not match_type:
                        match_type = "дубль"
                        matched_pair = c2
                elif (c1["clean_title"] in c2["clean_title"] or c2["clean_title"] in c1["clean_title"]) and c1["clean_title"] and c2["clean_title"]:
                    matched_ids.append(c2["raw_id"])
                    if not match_type:
                        match_type = "вложенный"
                        matched_pair = c2
                elif len(c1["clean_desc_words"].intersection(c2["clean_desc_words"])) >= 2:
                    matched_ids.append(c2["raw_id"])
                    if not match_type:
                        match_type = "вложенный"
                        matched_pair = c2

            if matched_pair:
                q1 = self.extract_quotes_content(c1["raw_title"])
                q2 = self.extract_quotes_content(matched_pair["raw_title"])
                if q1 and q2 and q1 != q2:
                    status = "Требует внимания"
                    comment = "Разные кавычки"
                auds1, lvl1 = self.get_audience_and_levels(c1["raw_title"])
                auds2, lvl2 = self.get_audience_and_levels(matched_pair["raw_title"])
                if auds1 and auds2 and auds1 != auds2:
                    status = "Требует внимания"
                    comment = "Разная ЦА"
                elif bool(auds1) != bool(auds2):
                    status = "Требует внимания"
                    comment = "Разная ЦА"
                if lvl1 and lvl2 and lvl1 != lvl2:
                    status = "Требует внимания"
                    comment = "Разный уровень сложности"

                matches.append({
                    "id": c1["raw_id"],
                    "title": c1["raw_title"],
                    "match_type": match_type,
                    "status": status,
                    "similar_ids": ", ".join(sorted(list(set(matched_ids)))),
                    "comment": comment
                })

        report_name = f"Отчет_по_дублям_каталога_{get_current_date_str()}.xlsx"
        report_path = os.path.join(self.workspace_dir, report_name)
        report_rows = [[i+1, m["id"], m["title"], m["match_type"], m["status"], m["similar_ids"], m["comment"]] for i, m in enumerate(matches)]
        hdr = ["№", "ID курса", "Название", "Тип совпадения", "Статус", "ID схожих курсов", "Комментарий"]

        if pd and openpyxl:
            pd.DataFrame(report_rows, columns=hdr).to_excel(report_path, index=False)
        elif openpyxl:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Свод дублей"
            ws.append(hdr)
            for r in report_rows:
                ws.append(r)
            wb.save(report_path)
        else:
            report_path = report_path.replace(".xlsx", ".csv")
            with open(report_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(hdr)
                writer.writerows(report_rows)

        total_courses = len(courses)
        total_matches = len(matches)
        duplicates_count = sum(1 for m in matches if m["match_type"] == "дубль")
        nested_count = sum(1 for m in matches if m["match_type"] == "вложенный")
        attention_count = sum(1 for m in matches if m["status"] == "Требует внимания")

        summary_text = (
            f"ОТЧЁТ ПО АНАЛИЗУ ДУБЛЕЙ КАТАЛОГА КУРСОВ ({get_current_date_str()})\n"
            f"========================================================\n"
            f"Всего курсов проанализировано: {total_courses}\n"
            f"Найдено пересечений/совпадений: {total_matches}\n"
            f"  • Из них точных дублей (по названию): {duplicates_count}\n"
            f"  • Из них частичных/вложенных совпадений: {nested_count}\n"
            f"  • Требуют оперативного внимания (стоп-фильтры): {attention_count}\n"
            f"Сводная таблица сохранена в файл: {os.path.basename(report_path)}\n"
        )
        return {"summary": summary_text, "report_file": report_path}


# =====================================================================
# AGENT 2: PII / PERSONAL DATA ANALYZER
# =====================================================================

class PIIAnalyzerAgent:
    def __init__(self, workspace_dir):
        self.workspace_dir = workspace_dir
        self.surnames = {"иванов", "петров", "смирнов", "сидоров", "кузнецов", "соколов", "васильев", "новиков", "попов"}
        self.names = {"александр", "екатерина", "саша", "катя", "иван", "петр", "сергей", "дмитрий", "алексей", "анна", "елена", "мария", "ольга", "татьяна"}
        self.ignore_preceding_markers = {"памятник", "кейс", "пример", "образец", "тестовый", "учебный"}

    def luhn_checksum_valid(self, card_number):
        digits = [int(x) for x in card_number if x.isdigit()]
        if len(digits) < 13:
            return False
        odd_digits = digits[-1::-2]
        even_digits = digits[-2::-2]
        total_sum = sum(odd_digits)
        for d in even_digits:
            doubled = d * 2
            total_sum += doubled if doubled < 10 else doubled - 9
        return total_sum % 10 == 0

    def parse_card_pattern(self, text):
        cards_found = []
        pattern = r'\b[245]\d{3}[-\s]?\d{4}[-\s]?\d{4}[-\s]?\d{4}\b'
        for m in re.finditer(pattern, text):
            snippet = m.group(0)
            cleaned = snippet.replace(" ", "").replace("-", "")
            if cleaned == "4111111111111111":
                continue
            if self.luhn_checksum_valid(cleaned):
                cards_found.append((snippet, cleaned))
        return cards_found

    def parse_phone_pattern(self, text):
        phones_found = []
        pattern = r'\+?[\d][-\s\(]*\d{3}[-\s\)]*\d{3}[-\s]*\d{2}[-\s]*\d{2}\b'
        for m in re.finditer(pattern, text):
            raw_snippet = m.group(0)
            cleaned = re.sub(r'[^\dXxХх]', '', raw_snippet)
            if len(cleaned) in (10, 11):
                phones_found.append((raw_snippet, cleaned))
        return phones_found

    def parse_fio_pattern(self, text):
        fios_found = []
        words = text.split()
        for idx, word in enumerate(words):
            cleaned_word = re.sub(r'[^\w.-]', '', word).strip()
            word_lower = cleaned_word.lower()
            if word_lower in self.surnames or word_lower in self.names or word_lower.endswith(("ич", "вна", "овна")):
                is_ignored = False
                for back_j in range(max(0, idx-2), idx):
                    if words[back_j].lower().strip(".,:;!?()\"'") in self.ignore_preceding_markers:
                        is_ignored = True
                        break
                if is_ignored:
                    continue
                seq_words = [words[idx + step] for step in range(3) if idx + step < len(words)]
                fios_found.append(" ".join(seq_words))
        return fios_found

    def analyze(self):
        possible_files = glob.glob(os.path.join(self.workspace_dir, "landing_text.*"))
        if not possible_files:
            possible_files = glob.glob(os.path.join(self.workspace_dir, "**", "landing_text.*"), recursive=True)
        if not possible_files:
            possible_files = glob.glob(os.path.join(self.workspace_dir, "**", "*риторика*"), recursive=True)
        if not possible_files:
            possible_files = glob.glob(os.path.join(self.workspace_dir, "**", "ЛЕНДИНГ.docx"), recursive=True)
        if not possible_files:
            raise FileNotFoundError(f"Landing text file not found under '{self.workspace_dir}'")

        file_path = possible_files[0]
        print(f"[PIIAnalyzerAgent] Reading document: {file_path}")
        raw_text = read_text_document(file_path)

        phones = self.parse_phone_pattern(raw_text)
        fios = self.parse_fio_pattern(raw_text)
        cards = self.parse_card_pattern(raw_text)

        report_rows = []
        for p_snippet, p_clean in phones:
            report_rows.append([p_clean, "Телефон", f"«... {p_snippet} ...»", "Удалить/заменить/замаскировать"])
        for f_val in fios:
            report_rows.append([f_val, "ФИО", f"«... {f_val} ...»", "Заменить"])
        for c_snippet, c_clean in cards:
            report_rows.append([c_clean, "Карта", f"«... {c_snippet} ...»", "Удалить или замаскировать"])

        report_name = f"Отчет_по_ПДн_документы_{get_current_date_str()}.xlsx"
        report_path = os.path.join(self.workspace_dir, report_name)
        hdr = ["Что нашёл", "Тип", "Где именно (фрагмент текста)", "Что делать"]

        if pd and openpyxl:
            pd.DataFrame(report_rows, columns=hdr).to_excel(report_path, index=False)
        elif openpyxl:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "ПДн Утечки"
            ws.append(hdr)
            for r in report_rows:
                ws.append(r)
            wb.save(report_path)
        else:
            report_path = report_path.replace(".xlsx", ".csv")
            with open(report_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(hdr)
                writer.writerows(report_rows)

        summary_text = (
            f"ОТЧЁТ ПО АНАЛИЗУ ПЕРСОНАЛЬНЫХ ДАННЫХ (ПДн) ({get_current_date_str()})\n"
            f"====================================================\n"
            f"Всего проверено документов: 1\n"
            f"Найдено телефонных номеров: {len(phones)}\n"
            f"Найдено совпадений ФИО: {len(fios)}\n"
            f"Найдено номеров карт: {len(cards)}\n"
            f"Детальный файл уязвимостей сохранен в: {os.path.basename(report_path)}\n"
        )
        return {"summary": summary_text, "report_file": report_path}


# =====================================================================
# AGENT 3: COMMENTS ANALYZER
# =====================================================================

class CommentsAnalyzerAgent:
    def __init__(self, workspace_dir):
        self.workspace_dir = workspace_dir
        self.bug_markers = ["не грузит", "вылетает", "ошибка", "завис", "не открывается", "белый экран", "кнопка не работает"]
        self.pain_markers = ["сложно", "непонятно", "мало практики", "много воды", "нелогично", "противоречие", "скучно"]

    def clean_comment(self, text):
        if not text or not isinstance(text, str):
            return ""
        text = text.encode('ascii', 'ignore').decode('ascii', errors='ignore')
        return " ".join(text.split())

    def analyze(self, file_name="Комментарии по курсам_синтетика.xlsx"):
        file_path = os.path.join(self.workspace_dir, file_name)
        if not os.path.exists(file_path):
            alternatives = glob.glob(os.path.join(self.workspace_dir, "**", file_name), recursive=True)
            if alternatives:
                file_path = alternatives[0]
            else:
                raise FileNotFoundError(f"Input file '{file_name}' not found.")

        print(f"[CommentsAnalyzerAgent] Analytical parser reading: {file_path}")

        if pd:
            df = pd.read_excel(file_path)
        elif openpyxl:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            hdr = rows[0]
            df = {col: [] for col in hdr if col}
            for row in rows[1:]:
                for idx, val in enumerate(row):
                    if idx < len(hdr) and hdr[idx]:
                        df[hdr[idx]].append(val)
        else:
            raise ImportError("Neither pandas nor openpyxl is available.")

        keys = list(df.keys())
        comment_col = next((c for c in keys if "коммент" in c.lower()), keys[0])
        id_col = next((c for c in keys if "id" in c.lower() or "№" in c.lower() or "номер" in c.lower()), keys[0])
        content_id_col = next((c for c in keys if "id контента" in c.lower() or "id_контента" in c.lower()), None)

        comments_list = []
        num_rows = len(df[comment_col])
        for i in range(num_rows):
            cid = df[id_col][i] if id_col in df else f"C-{i+1}"
            raw_text = df[comment_col][i]
            content_id = df[content_id_col][i] if content_id_col else None
            if not raw_text or str(raw_text).strip() == "nan":
                continue
            comments_list.append({
                "id": str(cid),
                "text": str(raw_text).strip(),
                "content_id": str(content_id).strip() if content_id else None
            })

        print(f"[CommentsAnalyzerAgent] Processing {len(comments_list)} feedback rows...")

        classified = []
        bugs = []
        pains = []
        positive_count = 0
        negative_count = 0
        neutral_count = 0

        for c in comments_list:
            text_lower = c["text"].lower()
            cleaned_text = self.clean_comment(c["text"])
            root = "unclear"
            confidence = 0.5

            is_bug = False
            bug_match = None
            for m in self.bug_markers:
                if m in text_lower:
                    is_bug = True
                    bug_match = m
                    break

            if is_bug:
                bugs.append({
                    "id": c["id"],
                    "место_бага": "плеер/навигация" if any(w in text_lower for w in ["грузит", "видео", "плеер"]) else "интерфейс",
                    "действие_пользователя": "нажатие кнопки" if "кнопка" in text_lower else "просмотр",
                    "что_увидел": bug_match,
                    "id_контента": c["content_id"] if c["content_id"] else "N/A"
                })
                root = "bug"
                confidence = 0.95

            is_pain = False
            for m in self.pain_markers:
                if m in text_lower:
                    is_pain = True
                    pains.append({
                        "id": c["id"],
                        "категория_боли": "сложность" if "сложно" in text_lower else "недостаток практики",
                        "объект": "Тема курса",
                        "суть": f"Пользователь жалуется на {m}",
                        "id_контента": c["content_id"] if c["content_id"] else "N/A"
                    })
                    break

            if is_pain and not is_bug:
                root = "pain"
                confidence = 0.90

            if any(p in text_lower for p in ["супер", "отлично", "рекомендую"]):
                sentiment = "positive"
                positive_count += 1
            elif any(n in text_lower for n in ["ужасно", "разочарован", "не советую"]):
                sentiment = "negative"
                negative_count += 1
            else:
                sentiment = "neutral"
                neutral_count += 1

            if not is_bug and not is_pain:
                root = "sentiment_only" if sentiment != "neutral" else "unclear"
                confidence = 0.70 if sentiment != "neutral" else 0.50

            classified.append({
                "id": c["id"],
                "text": c["text"],
                "cleaned": cleaned_text,
                "root_cause": root,
                "sentiment": sentiment,
                "confidence": confidence,
                "content_id": c["content_id"]
            })

        total = len(classified) if classified else 1
        pos_pct = (positive_count / total) * 100
        neg_pct = (negative_count / total) * 100
        neu_pct = (neutral_count / total) * 100
        prob_coeff = (len(bugs) + len(pains)) / total

        top_bugs_str = "\n".join(
            f"    {i}. {b['место_бага']} — {b['действие_пользователя']} — {b['что_увидел']} (id контента: {b['id_контента']}) (встречается 1 раз)"
            for i, b in enumerate(bugs[:3], 1)
        ) or "    (не найдено)"

        top_pains_str = "\n".join(
            f"    {i}. {p['объект']} — {p['категория_боли']} — {p['суть']} (id контента: {p['id_контента']}) (упоминаний 1)"
            for i, p in enumerate(pains[:3], 1)
        ) or "    (не найдено)"

        critical_list = [
            f"• [{x['id']}] Текст: \"{x['text']}\" (id контента: {x['content_id'] if x['content_id'] else 'N/A'})"
            for x in sorted(classified, key=lambda v: v["confidence"], reverse=True)
            if x["sentiment"] == "negative" or x["root_cause"] == "bug"
        ]
        critical_str = "\n".join(critical_list[:5]) if critical_list else "  • Опасные критические инциденты отсутствуют."

        report_text = (
            f"ОТЧЁТ ПО АНАЛИЗУ КОММЕНТАРИЕВ К КУРСУ\n"
            f"Дата: {get_current_date_str()}\n"
            f"Всего обработано: {total} комментариев\n\n"
            f"1. ОБЩАЯ ТОНАЛЬНОСТЬ\n"
            f"  • Позитивных: {positive_count} ({pos_pct:.1f}%)\n"
            f"  • Негативных: {negative_count} ({neg_pct:.1f}%)\n"
            f"  • Нейтральных: {neutral_count} ({neu_pct:.1f}%)\n\n"
            f"2. ТЕХНИЧЕСКИЕ БАГИ (приоритетно)\n"
            f"  ➜ Топ-3 бага:\n{top_bugs_str}\n"
            f"  ➜ Всего уникальных багов: {len(bugs)}\n\n"
            f"3. ПРОБЛЕМЫ КОНТЕНТА И МЕТОДИКИ (БОЛИ)\n"
            f"  ➜ Топ-3 боли:\n{top_pains_str}\n"
            f"  ➜ Всего уникальных болей: {len(pains)}\n\n"
            f"4. КРИТИЧЕСКИЕ КОММЕНТАРИИ (требуют немедленного внимания)\n"
            f"{critical_str}\n\n"
            f"5. КОЭФФИЦИЕНТ ПРОБЛЕМНОСТИ\n"
            f"  • Общий балл: {prob_coeff:.2f}\n\n"
            f"6. РЕКОМЕНДАЦИИ\n"
            f"  • По багам: проверить места возникновения багов в навигации, обновить медиа-компоненты.\n"
            f"  • По болям: пересмотреть сложные темы, добавить практические упражнения.\n"
        )

        return {"summary": report_text, "report_file": file_path}


# =====================================================================
# AGENT 4: POSTMAN - LOCAL ONLY (.eml files)
# =====================================================================

class PostmanAgent:
    def __init__(self):
        self.save_directory = "/Users/uliacibisova/Desktop/Результат"

    def save_email(self, recipient, subject, body_text, attachment_paths=None):
        msg = MIMEMultipart()
        msg['From'] = "Ouroboros Agent <ouroboros@agent.local>"
        msg['To'] = recipient
        msg['Subject'] = subject
        msg.attach(MIMEText(body_text, 'plain', 'utf-8'))

        if attachment_paths:
            for path in attachment_paths:
                if os.path.exists(path):
                    with open(path, "rb") as f:
                        part = MIMEApplication(f.read(), Name=os.path.basename(path))
                    part['Content-Disposition'] = f'attachment; filename="{os.path.basename(path)}"'
                    msg.attach(part)

        os.makedirs(self.save_directory, exist_ok=True)
        sanitized_subject = re.sub(r'[^\w\s-]', '', subject.replace(' ', '_'))
        eml_filename = f"Mail_Digest_{sanitized_subject}_{get_current_date_str()}.eml"
        eml_path = os.path.join(self.save_directory, eml_filename)

        with open(eml_path, "w", encoding="utf-8") as f:
            f.write(msg.as_string())

        print(f"[PostmanAgent] Saved: {os.path.abspath(eml_path)}")
        return eml_path


# =====================================================================
# SYSTEM ORCHESTRATOR
# =====================================================================

class SystemOrchestrator:
    def __init__(self, workspace_dir):
        self.workspace_dir = workspace_dir
        self.duplicates_agent = DuplicatesAnalyzerAgent(workspace_dir)
        self.pii_agent = PIIAnalyzerAgent(workspace_dir)
        self.comments_agent = CommentsAnalyzerAgent(workspace_dir)
        self.postman_agent = PostmanAgent()

    def run_entire_pipeline(self, email_recipient="olkalozenko@gmail.com"):
        print("\n=== STARTING MULTI-AGENT ANALYTICAL EXECUTION ===\n")

        # 1. Duplicates Analysis
        try:
            res_dup = self.duplicates_agent.analyze()
            self.postman_agent.save_email(
                recipient=email_recipient,
                subject=f"Отчет по дублям от {get_current_date_str()}",
                body_text=res_dup["summary"],
                attachment_paths=[res_dup["report_file"]]
            )
            print(res_dup["summary"])
        except Exception as e:
            print(f"[Orchestrator] Ошибка в Агенте 1 (Дубли): {e}")

        # 2. PII / Personal Data Analysis
        try:
            res_pii = self.pii_agent.analyze()
            self.postman_agent.save_email(
                recipient=email_recipient,
                subject=f"Отчет по ПДн от {get_current_date_str()}",
                body_text=res_pii["summary"],
                attachment_paths=[res_pii["report_file"]]
            )
            print(res_pii["summary"])
        except Exception as e:
            print(f"[Orchestrator] Ошибка в Агенте 2 (ПДн): {e}")

        # 3. Comments Analysis
        try:
            res_comm = self.comments_agent.analyze()
            self.postman_agent.save_email(
                recipient=email_recipient,
                subject=f"Отчет по комментариям от {get_current_date_str()}",
                body_text=res_comm["summary"],
                attachment_paths=[res_comm["report_file"]]
            )
            print(res_comm["summary"])
        except Exception as e:
            print(f"[Orchestrator] Ошибка в Агенте 3 (Комментарии): {e}")

        print("=== MULTI-AGENT EXECUTION COMPLETE ===\n")
        return "Конвейер завершён. Отчёты сохранены в папку Результат на Рабочем столе."


if __name__ == "__main__":
    workspace = "/Users/uliacibisova/Desktop/Audit"
    orchestrator = SystemOrchestrator(workspace)
    digest = orchestrator.run_entire_pipeline()
    print(digest)
